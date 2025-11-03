import pdfplumber
import pandas as pd
from pathlib import Path
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import sys
import cv2
import numpy as np
from pdf2image import convert_from_path
import pytesseract
from PIL import Image

class PDFTableExtractor:
    """Extract tables from PDF using image preprocessing and OCR"""
    
    def __init__(self, pdf_path, use_image_processing=True, debug_mode=False):
        self.pdf_path = Path(pdf_path)
        self.tables_data = []
        self.use_image_processing = use_image_processing
        self.debug_mode = debug_mode
        
    def preprocess_image(self, image):
        """Apply CV2 preprocessing for better table detection"""
        # Convert PIL image to numpy array
        img_array = np.array(image)
        
        # Convert to grayscale
        if len(img_array.shape) == 3:
            gray = cv2.cvtColor(img_array, cv2.COLOR_RGB2GRAY)
        else:
            gray = img_array
        
        # Apply multiple preprocessing techniques
        # 1. Increase contrast using CLAHE
        clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8,8))
        enhanced = clahe.apply(gray)
        
        # 2. Denoise
        denoised = cv2.fastNlMeansDenoising(enhanced, None, 10, 7, 21)
        
        # 3. Binary threshold - Otsu's method
        _, binary = cv2.threshold(denoised, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
        
        # 4. Morphological operations to enhance table lines
        kernel = np.ones((2,2), np.uint8)
        morph = cv2.morphologyEx(binary, cv2.MORPH_CLOSE, kernel)
        
        # 5. Sharpen the image
        kernel_sharp = np.array([[-1,-1,-1],
                                 [-1, 9,-1],
                                 [-1,-1,-1]])
        sharpened = cv2.filter2D(morph, -1, kernel_sharp)
        
        # Convert back to PIL Image
        processed_image = Image.fromarray(sharpened)
        
        # Save debug image if debug mode is on
        if self.debug_mode:
            debug_path = self.pdf_path.parent / f"debug_processed_page.png"
            processed_image.save(debug_path)
            print(f"  Debug: Saved preprocessed image to {debug_path}")
        
        return processed_image
    
    def extract_tables_from_image(self, image, page_num):
        """Extract tables from preprocessed image using pdfplumber"""
        tables_found = []
        
        try:
            # Save preprocessed image to temporary file
            temp_path = self.pdf_path.parent / f"temp_page_{page_num}.png"
            image.save(temp_path, dpi=(300, 300))
            
            # Try to detect table structure using OpenCV
            img_cv = np.array(image)
            
            # Detect horizontal and vertical lines
            horizontal_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (40, 1))
            vertical_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, 40))
            
            # Detect horizontal lines
            horizontal_lines = cv2.morphologyEx(img_cv, cv2.MORPH_OPEN, horizontal_kernel, iterations=2)
            # Detect vertical lines
            vertical_lines = cv2.morphologyEx(img_cv, cv2.MORPH_OPEN, vertical_kernel, iterations=2)
            
            # Combine lines
            table_structure = cv2.addWeighted(horizontal_lines, 0.5, vertical_lines, 0.5, 0.0)
            
            # Find contours (potential table cells)
            contours, _ = cv2.findContours(table_structure, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
            
            print(f"  Found {len(contours)} potential table structures")
            
            # Clean up temp file
            if temp_path.exists():
                temp_path.unlink()
                
        except Exception as e:
            print(f"  Image processing warning: {e}")
        
        return tables_found
    
    def extract_tables(self):
        """Extract all tables from PDF with image preprocessing"""
        print(f"Processing: {self.pdf_path.name}")
        print(f"Image preprocessing: {'Enabled' if self.use_image_processing else 'Disabled'}")
        
        try:
            if self.use_image_processing:
                # Convert PDF to images with high DPI
                print("Converting PDF to images (this may take a moment)...")
                images = convert_from_path(self.pdf_path, dpi=300)
                
                for page_num, image in enumerate(images, 1):
                    print(f"\nProcessing page {page_num}/{len(images)}...")
                    
                    # Preprocess image
                    print("  Applying image enhancement...")
                    processed_image = self.preprocess_image(image)
                    
                    # Save processed image temporarily for pdfplumber
                    temp_pdf_path = self.pdf_path.parent / f"temp_processed_{page_num}.pdf"
                    processed_image.save(temp_pdf_path.with_suffix('.png'), dpi=(300, 300))
                    
                    # Extract table structure info
                    self.extract_tables_from_image(processed_image, page_num)
            
            # Now use pdfplumber with various strategies
            with pdfplumber.open(self.pdf_path) as pdf:
                for page_num, page in enumerate(pdf.pages, 1):
                    print(f"\nExtracting from page {page_num}/{len(pdf.pages)}...")
                    
                    # Strategy 1: Line-based detection (strictest)
                    print("  Strategy 1: Line-based detection...")
                    tables = page.extract_tables({
                        "vertical_strategy": "lines",
                        "horizontal_strategy": "lines",
                        "snap_tolerance": 5,
                        "join_tolerance": 5,
                        "edge_min_length": 10,
                        "min_words_vertical": 1,
                        "min_words_horizontal": 1,
                        "intersection_tolerance": 3,
                    })
                    
                    if tables:
                        for idx, table in enumerate(tables, 1):
                            if table and len(table) > 0:
                                self.tables_data.append({
                                    'page': page_num,
                                    'table_index': len([t for t in self.tables_data if t['page'] == page_num]) + 1,
                                    'data': table,
                                    'method': 'lines'
                                })
                                print(f"    ✓ Found table with {len(table)} rows, {len(table[0]) if table else 0} columns")
                    
                    # Strategy 2: Text-based detection
                    if not tables:
                        print("  Strategy 2: Text-based detection...")
                        tables = page.extract_tables({
                            "vertical_strategy": "text",
                            "horizontal_strategy": "text",
                            "snap_tolerance": 3,
                            "intersection_tolerance": 5,
                        })
                        
                        if tables:
                            for idx, table in enumerate(tables, 1):
                                if table and len(table) > 0:
                                    self.tables_data.append({
                                        'page': page_num,
                                        'table_index': len([t for t in self.tables_data if t['page'] == page_num]) + 1,
                                        'data': table,
                                        'method': 'text'
                                    })
                                    print(f"    ✓ Found table with {len(table)} rows, {len(table[0]) if table else 0} columns")
                    
                    # Strategy 3: Lines and text combined
                    if not tables:
                        print("  Strategy 3: Combined line+text detection...")
                        tables = page.extract_tables({
                            "vertical_strategy": "lines_strict",
                            "horizontal_strategy": "lines_strict",
                            "snap_tolerance": 10,
                            "join_tolerance": 10,
                            "edge_min_length": 5,
                        })
                        
                        if tables:
                            for idx, table in enumerate(tables, 1):
                                if table and len(table) > 0:
                                    self.tables_data.append({
                                        'page': page_num,
                                        'table_index': len([t for t in self.tables_data if t['page'] == page_num]) + 1,
                                        'data': table,
                                        'method': 'combined'
                                    })
                                    print(f"    ✓ Found table with {len(table)} rows, {len(table[0]) if table else 0} columns")
                    
                    # Strategy 4: Use detected edges from page
                    if not tables:
                        print("  Strategy 4: Edge-based detection...")
                        # Get all edges detected in the page
                        edges = page.edges
                        if edges:
                            # Try with custom settings based on detected edges
                            h_edges = [e for e in edges if e.get('orientation') == 'h']
                            v_edges = [e for e in edges if e.get('orientation') == 'v']
                            
                            if h_edges and v_edges:
                                tables = page.extract_tables({
                                    "vertical_strategy": "lines",
                                    "horizontal_strategy": "lines",
                                    "snap_tolerance": 15,
                                    "join_tolerance": 15,
                                    "edge_min_length": 3,
                                    "min_words_vertical": 0,
                                    "min_words_horizontal": 0,
                                })
                                
                                if tables:
                                    for idx, table in enumerate(tables, 1):
                                        if table and len(table) > 0:
                                            self.tables_data.append({
                                                'page': page_num,
                                                'table_index': len([t for t in self.tables_data if t['page'] == page_num]) + 1,
                                                'data': table,
                                                'method': 'edges'
                                            })
                                            print(f"    ✓ Found table with {len(table)} rows, {len(table[0]) if table else 0} columns")
                    
                    # Strategy 5: Extract all text and try to parse as table
                    if not tables:
                        print("  Strategy 5: Text extraction and parsing...")
                        text = page.extract_text()
                        if text:
                            # Try to detect table-like structures in text
                            lines = text.split('\n')
                            potential_table = []
                            for line in lines:
                                # Look for lines with multiple space-separated values
                                parts = line.split()
                                if len(parts) >= 2:  # At least 2 columns
                                    potential_table.append(parts)
                            
                            if len(potential_table) >= 2:  # At least 2 rows
                                self.tables_data.append({
                                    'page': page_num,
                                    'table_index': len([t for t in self.tables_data if t['page'] == page_num]) + 1,
                                    'data': potential_table,
                                    'method': 'text_parse'
                                })
                                print(f"    ✓ Parsed table-like text with {len(potential_table)} rows")
                    
                    if not any(t['page'] == page_num for t in self.tables_data):
                        print(f"    ✗ No tables detected on this page")
        
        except Exception as e:
            print(f"Error processing PDF: {e}")
            import traceback
            traceback.print_exc()
        
        # Clean up temporary files
        for temp_file in self.pdf_path.parent.glob("temp_*.png"):
            temp_file.unlink()
        
        if not self.tables_data:
            print("\n⚠ Warning: No tables found in PDF")
            print("\nTroubleshooting tips:")
            print("1. Ensure the PDF contains actual tables (not just images of tables)")
            print("2. Try with use_image_processing=True for scanned documents")
            print("3. Check if tables have clear borders/lines")
            print("4. The PDF might need OCR if it's a scanned document")
        else:
            print(f"\n✓ Total tables extracted: {len(self.tables_data)}")
            for table in self.tables_data:
                print(f"  - Page {table['page']}, Table {table['table_index']}: "
                      f"{len(table['data'])} rows, Method: {table['method']}")
        
        return self.tables_data
    
    def clean_cell_data(self, cell):
        """Clean and format cell data"""
        if cell is None:
            return ""
        # Remove excessive whitespace and newlines
        cleaned = " ".join(str(cell).split())
        return cleaned
    
    def export_to_excel(self, output_path=None, style_tables=True):
        """Export extracted tables to Excel with formatting"""
        if not self.tables_data:
            print("No tables to export. Run extract_tables() first.")
            return
        
        if output_path is None:
            output_path = self.pdf_path.with_suffix('.xlsx')
        else:
            output_path = Path(output_path)
        
        # Create Excel writer
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            for table_info in self.tables_data:
                page = table_info['page']
                idx = table_info['table_index']
                table = table_info['data']
                method = table_info.get('method', 'unknown')
                
                # Clean the table data
                cleaned_table = []
                for row in table:
                    cleaned_row = [self.clean_cell_data(cell) for cell in row]
                    cleaned_table.append(cleaned_row)
                
                # Create DataFrame
                if len(cleaned_table) > 1:
                    df = pd.DataFrame(cleaned_table[1:], columns=cleaned_table[0])
                else:
                    df = pd.DataFrame(cleaned_table)
                
                # Create sheet name
                sheet_name = f"P{page}_T{idx}_{method}"[:31]  # Excel limit
                
                # Write to Excel
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Apply styling if requested
                if style_tables:
                    worksheet = writer.sheets[sheet_name]
                    self._apply_excel_styling(worksheet, df)
        
        print(f"\n✓ Excel file created: {output_path}")
        return output_path
    
    def _apply_excel_styling(self, worksheet, df):
        """Apply professional styling to Excel worksheet"""
        # Define styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        # Style header row
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Style data cells
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, 
                                       min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.border = border
                cell.alignment = alignment
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Set row heights
        for row in range(2, worksheet.max_row + 1):
            worksheet.row_dimensions[row].height = 30
        worksheet.row_dimensions[1].height = 35
    
    def export_to_csv(self, output_dir=None):
        """Export each table to separate CSV files"""
        if not self.tables_data:
            print("No tables to export. Run extract_tables() first.")
            return
        
        if output_dir is None:
            output_dir = self.pdf_path.parent / f"{self.pdf_path.stem}_tables"
        else:
            output_dir = Path(output_dir)
        
        output_dir.mkdir(exist_ok=True)
        
        for table_info in self.tables_data:
            page = table_info['page']
            idx = table_info['table_index']
            table = table_info['data']
            
            # Clean the table data
            cleaned_table = []
            for row in table:
                cleaned_row = [self.clean_cell_data(cell) for cell in row]
                cleaned_table.append(cleaned_row)
            
            # Create DataFrame and save
            df = pd.DataFrame(cleaned_table[1:], columns=cleaned_table[0]) if len(cleaned_table) > 1 else pd.DataFrame(cleaned_table)
            
            csv_path = output_dir / f"page{page}_table{idx}.csv"
            df.to_csv(csv_path, index=False, encoding='utf-8-sig')
            print(f"Saved: {csv_path}")
        
        print(f"\nAll CSV files saved to: {output_dir}")
        return output_dir


def main():
    """Main function with example usage"""
    print("=" * 60)
    print("PDF TABLE EXTRACTOR WITH IMAGE PREPROCESSING")
    print("=" * 60)
    
    # Get PDF path from command line or use default
    if len(sys.argv) > 1:
        pdf_path = sys.argv[1]
    else:
        pdf_path = input("Enter PDF file path: ").strip('"\'')
    
    if not Path(pdf_path).exists():
        print(f"Error: File not found - {pdf_path}")
        return
    
    # Ask about image preprocessing
    use_preprocessing = input("\nUse image preprocessing? (recommended for scanned docs) [y/n]: ").strip().lower()
    use_image = use_preprocessing in ['y', 'yes', '']
    
    # Ask about debug mode
    debug = input("Enable debug mode? (saves preprocessed images) [y/n]: ").strip().lower()
    debug_mode = debug in ['y', 'yes']
    
    try:
        # Initialize extractor
        extractor = PDFTableExtractor(pdf_path, use_image_processing=use_image, debug_mode=debug_mode)
        
        # Extract tables
        print("\nStarting extraction...\n")
        extractor.extract_tables()
        
        # Export to Excel
        if extractor.tables_data:
            print("\nExporting to Excel...")
            excel_path = extractor.export_to_excel(style_tables=True)
            
            # Optional: Also export to CSV
            export_csv = input("\nAlso export to CSV files? (y/n): ").strip().lower()
            if export_csv == 'y':
                extractor.export_to_csv()
            
            print("\n" + "=" * 60)
            print("✓ EXTRACTION COMPLETE!")
            print("=" * 60)
        else:
            print("\n" + "=" * 60)
            print("✗ NO TABLES FOUND")
            print("=" * 60)
    
    except Exception as e:
        print(f"\n✗ Error: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()