import os
from main import save_equipment_to_excel
import shutil

# Mock Data that matches the new prompt structure
mock_data = {
    "equipment_tag": "V-101",
    "description": "Scrubber Vessel",
    "pmt_no": "PMT-001",
    "equipment_type": "Pressure Vessel",
    "conditions": {
        "design_pressure": 1.5,
        "design_temp": 120,
        "operating_pressure": 1.2,
        "operating_temp": 110,
        "insulation": True,
    },
    "components": {
        "top_head": {
            "phase": "Vapor",
            "fluid": "Gas",
            "material": {"type": "CS", "spec": "SA-516", "grade": "70"},
        },
        "shell": {
            "phase": "Two-Phase",
            "fluid": "Mixed",
            "material": {"type": "CS", "spec": "SA-516", "grade": "70"},
        },
        "bottom_head": {
            "phase": "Liquid",
            "fluid": "Liquid",
            "material": {"type": "CS", "spec": "SA-516", "grade": "70"},
        },
    },
}

from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws["A5"] = "Placeholder Header"  # Just to make sure we skip to row 6
wb.save("test_template.xlsx")

output_file = "test_output.xlsx"
if os.path.exists(output_file):
    os.remove(output_file)

try:
    save_equipment_to_excel(mock_data, output_file, template_path="test_template.xlsx")
    print(f"✅ Executed save_equipment_to_excel successfully. Output: {output_file}")

    # Verify Content
    from openpyxl import load_workbook

    wb_out = load_workbook(output_file)
    ws_out = wb_out.active

    # Check Merges (Row 6 to 8)
    # verify that merges exist
    print(f"Merged cells: {ws_out.merged_cells.ranges}")

    assert len(ws_out.merged_cells.ranges) >= 1, "Should have merges!"

except Exception as e:
    print(f"❌ Verification Failed: {e}")
    import traceback

    traceback.print_exc()

# Cleanup
if os.path.exists("test_template.xlsx"):
    os.remove("test_template.xlsx")
