import sys
import os
import requests

# Ensure we can import from backend if needed, but we will test the endpoint via requests if server is running,
# or direct function call if we want unit test.
# Let's do direct function call to test logic first without relying on server state.

sys.path.append(os.getcwd())
from main import read_equipment_from_master_file, save_equipment_to_template_pptx

MASTER_FILE = "MasterFile_IPETRO_PLANT.xlsx"

if not os.path.exists(MASTER_FILE):
    print(f"Skipping test: {MASTER_FILE} does not exist in current dir.")
    # Assuming the test is run from backend dir where the file might be.
    # The actual file is in OUTPUT_FOLDER..
    # Let's check OUTPUT folder
    OUTPUT_FOLDER = "outputs"  # Based on main.py
    MASTER_PATH = os.path.join(OUTPUT_FOLDER, MASTER_FILE)
else:
    MASTER_PATH = MASTER_FILE

if not os.path.exists(MASTER_PATH):
    print(f"Error: {MASTER_PATH} not found. Cannot test reading.")
    sys.exit(1)

print(f"Testing reading from {MASTER_PATH}...")

# 1. Read 'V-001' (Assuming it exists from previous steps or user uploads)
# If not, let's try to find ANY tag
from openpyxl import load_workbook

wb = load_workbook(MASTER_PATH, data_only=True)
ws = wb.active
first_tag = ws.cell(row=6, column=2).value
if not first_tag:
    print("No data in Master File to test with.")
    sys.exit(0)

print(f"Found tag to test: {first_tag}")

data = read_equipment_from_master_file(str(first_tag), MASTER_PATH)
if not data:
    print("Failed to read data.")
    sys.exit(1)

print("Data read successfully:")
print(data)

# 2. Test PPTX Generation
print("Generating PPTX...")
test_pptx_out = "test_sync_gen.pptx"
save_equipment_to_template_pptx(data, test_pptx_out, image_path=None)

if os.path.exists(test_pptx_out):
    print(f"SUCCESS: Generated {test_pptx_out}")
else:
    print("FAILURE: PPTX not created.")
