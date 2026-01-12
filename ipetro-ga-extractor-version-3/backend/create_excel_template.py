import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def create_excel_template():
    """ create excel template with placeholder """

    wb = Workbook()
    ws = wb.active
    ws.title = "Masterfile"

    # header section 

    # title
    ws['A1'] = 'IPETRO PLANT'
    ws['A1'].font = Font(size=16, bold=True)

    ws['A2'] = 'MASTERFILE EQUIPMENT'
    ws['A2'].font = Font(size=14, bold=True)

    ws['A3'] = 'TOTAL NO. OF EQUIPMENT: {{TOTAL_COUNT}} EQUIPMENTS'
    ws['A3'].font = Font(size=12, bold=True)

    # column headers (row 5 )

    headers = [
        'NO.',
        'EQUIPMENT NO.',
        'PMT NO.',
        'EQUIPMENT DESCRIPTION',
        'PARTS',
        'PHASE',
        'FLUID',
        'MATERIAL TYPE',
        'MATERIAL SPEC',
        'MATERIAL GRADE',
        'INSULATION\n(Yes/No)',
        'DESIGN\nTEMP.\n(°C)',
        'DESIGN\nPRESSURE\n(Mpa)',
        'OPERATING\nTEMP.\n(°C)',
        'OPERATING\nPRESSURE\n(Mpa)',
    ]

    # Header styling
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    # data rows with placeholders (starting row 6)
    placeholder_rows = [
        # Row 6 - Equipment with 3 parts (Pressure Vessel example)
        ['{{NO}}', '{{EQUIPMENT_TAG}}', '{{PMT_NO}}', '{{DESCRIPTION}}', 'Top Head', '{{PHASE_TOP}}', '{{FLUID_TOP}}', 
         '{{MATERIAL_TYPE_TOP}}', '{{MATERIAL_SPEC_TOP}}', '{{MATERIAL_GRADE_TOP}}', '{{INSULATION}}',
         '{{DESIGN_TEMP}}', '{{DESIGN_PRESSURE}}', '{{OPERATING_TEMP}}', '{{OPERATING_PRESSURE}}'],
        
        # Row 7 - Shell part (merged equipment info)
        ['', '', '', '', 'Shell', '{{PHASE_SHELL}}', '{{FLUID_SHELL}}',
         '{{MATERIAL_TYPE_SHELL}}', '{{MATERIAL_SPEC_SHELL}}', '{{MATERIAL_GRADE_SHELL}}', '',
         '{{DESIGN_TEMP}}', '{{DESIGN_PRESSURE}}', '{{OPERATING_TEMP}}', '{{OPERATING_PRESSURE}}'],
        
        # Row 8 - Bottom Head
        ['', '', '', '', 'Bottom Head', '{{PHASE_BOTTOM}}', '{{FLUID_BOTTOM}}',
         '{{MATERIAL_TYPE_BOTTOM}}', '{{MATERIAL_SPEC_BOTTOM}}', '{{MATERIAL_GRADE_BOTTOM}}', '',
         '{{DESIGN_TEMP}}', '{{DESIGN_PRESSURE}}', '{{OPERATING_TEMP}}', '{{OPERATING_PRESSURE}}'],
    ]

    # add placeholder rows
    start_row = 6
    for row_data in placeholder_rows:
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=start_row, column=col_num)
            cell.value = value
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        start_row += 1

    # auto fit columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    template_path = 'templates/equipment_template.xlsx'
    wb.save(template_path)
    print(f"Excel template created at: {template_path}")

    return template_path